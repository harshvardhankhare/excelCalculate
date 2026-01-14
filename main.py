from fastapi import FastAPI, HTTPException, BackgroundTasks
from pydantic import BaseModel, Field
from typing import List
import pandas as pd
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import uuid
import os

app = FastAPI(title="Loan Excel Generator API")

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # restrict in real production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------- INSURANCE LOGIC ----------------
def calculate_insurance_rate(loan_percentage: float, loan_period: int):
    if loan_percentage > 95.01:
        return None
    if 70 <= loan_percentage <= 80.99:
        return 0.0032
    if loan_percentage == 81:
        return 0.0021 if loan_period <= 25 else 0.0032
    if 81.01 <= loan_percentage <= 90:
        return 0.0041 if loan_period <= 25 else 0.0052
    if 90.01 <= loan_percentage <= 95:
        return 0.0067 if loan_period <= 25 else 0.0078
    return None

# ---------------- REQUEST MODEL ----------------
class LoanInput(BaseModel):
    sample_no: str
    customer_reference: str
    customer_name: str
    city_state: str

    A: float = Field(gt=0)
    down_payment: float = Field(ge=0, le=100)
    loan_period: int = Field(gt=0, le=40)
    annuity_interest: float = Field(ge=0, le=100)

    purchase_value_reduction: float = Field(ge=0, le=100)
    monthly_principal_reduction: float = Field(ge=0, le=100)
    total_interest_reduction: float = Field(ge=0, le=100)

    guarantor_name: str
    guarantor_reference: str

# ---------------- CALCULATION ----------------
def process_record(data: LoanInput):
    purchase_value = data.A * data.purchase_value_reduction / 100
    loan_amount = purchase_value * data.down_payment / 100

    base_principal = (loan_amount / data.loan_period) / 12
    principal = base_principal - (
        base_principal * data.monthly_principal_reduction / 100
    )

    base_interest = loan_amount * data.loan_period
    interest_value = base_interest * data.annuity_interest / 100
    total_interest = interest_value - (
        interest_value * data.total_interest_reduction / 100
    )

    loan_percentage = 100 - data.down_payment
    rate = calculate_insurance_rate(loan_percentage, data.loan_period)

    insurance_monthly = (
        "NA" if rate is None else round((loan_amount * rate) / 12, 2)
    )

    return {
        "Sample no,record no": data.sample_no,
        "Customer Reference Number": data.customer_reference,
        "Customer Name": data.customer_name,
        "City , State": data.city_state,
        "Purchase Value & Down Payment":
            f"$  {purchase_value:,.2f} and {data.down_payment}%",
        "Loan Period AND Annuity Interest":
            f"{data.loan_period} Years and {data.annuity_interest}%",
        "Guarantor Name": data.guarantor_name,
        "Guarantor Reference Number": data.guarantor_reference,
        "Loan Amount AND Principal":
            f"$  {loan_amount:,.2f} , {principal:,.2f}",
        "Total Interest for Loan":
            f"$  {total_interest:,.2f}",
        "Period & Property Insurance per Month":
            "NA" if insurance_monthly == "NA"
            else f"$  {insurance_monthly:,.2f}"
    }

# ---------------- HEALTH CHECK ----------------
@app.api_route("/ping", methods=["GET", "HEAD"])
def ping():
    return {"status": "alive"}

# ---------------- API ----------------
@app.post("/generate-excel")
def generate_excel(
    records: List[LoanInput],
    background_tasks: BackgroundTasks
):
    if not records:
        raise HTTPException(status_code=400, detail="No records provided")

    rows = [process_record(r) for r in records]
    df = pd.DataFrame(rows)

    file_name = f"/tmp/loan_{uuid.uuid4().hex}.xlsx"

    try:
        df.to_excel(file_name, index=False)

        wb = load_workbook(file_name)
        ws = wb.active

        # ----- COLUMN WIDTHS -----
        column_widths = {
            "A": 25.71,
            "B": 64.86,
            "C": 39.71,
            "D": 27.14,
            "E": 47.14,
            "F": 39.29,
            "G": 38.43,
            "H": 94.43,
            "I": 46.86,
            "J": 26.14
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # ----- ROW HEIGHT -----
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 18.75

        header_font = Font(bold=True, size=14)
        data_font = Font(size=14)

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        wb.save(file_name)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Auto-delete file after response
    background_tasks.add_task(os.remove, file_name)

    return FileResponse(
        path=file_name,
        filename="loan_calculation.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Start app:
# uvicorn main:app --host 0.0.0.0 --port 10000
