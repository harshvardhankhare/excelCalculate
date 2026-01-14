import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time


# ---------------------------------------------------
# INSURANCE RATE LOGIC (AS PROVIDED)
# ---------------------------------------------------
def calculate_insurance_rate(loan_percentage, loan_period):
    if loan_percentage > 95.01:
        return None

    if 70 <= loan_percentage <= 80.99:
        return 0.0032

    if loan_percentage == 81:
        return 0.0021 if loan_period <= 25 else 0.0032

    if 81.01 <= loan_percentage <= 90:
        return 0.0041 if loan_period <= 25 else 0.0052

    if 90.01 <= loan_percentage <= 95:
        return 0.0067 if loan_period <= 25 else 0.0078

    return None


# ---------------------------------------------------
# MAIN CALCULATION FUNCTION
# ---------------------------------------------------
def process_record(data):
    A = data["A"]

    # 1. Purchase Value = PVR % of A
    purchase_value = A * data["purchase_value_reduction"] / 100

    # 2. Loan Amount = Down Payment % of Purchase Value
    loan_amount = purchase_value * data["down_payment"] / 100

    # 3. Monthly Principal (reduced by %)
    base_principal = (loan_amount / data["loan_period"]) / 12
    principal = base_principal - (
        base_principal * data["monthly_principal_reduction"] / 100
    )

    # 4. Total Interest (exact order you defined)
    base_interest_value = loan_amount * data["loan_period"]
    interest_value = base_interest_value * data["annuity_interest"] / 100
    total_interest = interest_value - (
        interest_value * data["total_interest_reduction"] / 100
    )

    # 5. Loan Percentage
    loan_percentage = 100 - data["down_payment"]

    # 6. Property Insurance
    rate = calculate_insurance_rate(loan_percentage, data["loan_period"])

    if rate is None:
        insurance_monthly = "NA"
    else:
        insurance_per_annum = loan_amount * rate
        insurance_monthly = round(insurance_per_annum / 12, 2)

    # 7. Row for Excel
    return {
        "Sample No": data["sample_no"],
        "Customer Reference Number": data["customer_reference"],
        "Customer Name": data["customer_name"],
        "City State": data["city_state"],
        "Purchase Value & Down Payment":
            f"$  {purchase_value:,.2f} and {data['down_payment']}%",
        "Loan Period & Annuity Interest":
            f"{data['loan_period']} Years and {data['annuity_interest']}%",
        "Guarantor Name": data["guarantor_name"],
        "Guarantor Reference Number": data["guarantor_reference"],
        "Loan Amount & Principal":
            f"$  {loan_amount:,.2f} , {principal:,.2f}",
        "Total Interest for Loan":
            f"$  {total_interest:,.2f}",
        "Period & Property Insurance per Month":
            "NA" if insurance_monthly == "NA"
            else f"$  {insurance_monthly:,.2f}"
    }


# ---------------------------------------------------
# INPUT DATA
# ---------------------------------------------------
records = [
    {
        "sample_no": 1,
        "customer_reference": "CR12345",
        "customer_name": "John Doe",
        "city_state": "New York, NY",
        "A": 88850508.30,
        "down_payment": 29,
        "loan_period": 16,
        "annuity_interest": 8.7,
        "purchase_value_reduction": 14.56,
        "monthly_principal_reduction": 9.76,
        "total_interest_reduction": 15.42,
        "guarantor_name": "Mark Doe",
        "guarantor_reference": "GR98765"
    }
]


# ---------------------------------------------------
# CREATE EXCEL FILE (SAFE NAME)
# ---------------------------------------------------
rows = [process_record(r) for r in records]
df = pd.DataFrame(rows)

file_name = f"loan_calculation_{int(time.time())}.xlsx"
df.to_excel(file_name, index=False)


# ---------------------------------------------------
# EXCEL FORMATTING (BIG FONT + BIG CELLS)
# ---------------------------------------------------
wb = load_workbook(file_name)
ws = wb.active

# Column width
for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 34

# Row height
for r in range(1, ws.max_row + 1):
    ws.row_dimensions[r].height = 34

# Fonts
header_font = Font(bold=True, size=16)
data_font = Font(size=14)

# Header style
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

# Data cell style
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = data_font
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

# Freeze header row
ws.freeze_panes = "A2"

wb.save(file_name)

print("Excel created successfully with BIG FONT & BIG CELLS:")
print(file_name)
